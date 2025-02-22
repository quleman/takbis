from logging.handlers import RotatingFileHandler
import logging
import os
from datetime import datetime
import sys
import pdfplumber
import sqlite3
import re
import logging
from typing import Dict, Any, List, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal, getcontext, InvalidOperation  # InvalidOperation eklendi
import fitz  # PyMuPDF'in ana modülü
from typing import List, Dict, Optional  # Tip tanımlamaları için
import webbrowser

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTextEdit, QMessageBox, QFileDialog, QComboBox,
    QDialog, QProgressBar, QTreeWidgetItem, QSplitter, QTreeWidget, QMenu, QMenuBar, QAction)

from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont  # Bu satırı ekleyin

from takbisduzenle import TapuProcessor  # Yeni import
from ipotek_extractor import IpotekKoordinatExtractor
from takbis_inceleme import TakbisInceleme
from takbisler_inceleme import CokluInceleme  # Yeni import
from basliklar import FitzTapuAnalyzer

def dummy_log(*args, **kwargs):
    pass

logging.error = dummy_log
logging.info = dummy_log
logging.warning = dummy_log
logging.debug = dummy_log

# Log klasörü oluştur
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

# Log dosya adını tarih ile oluştur
log_filename = os.path.join(log_dir, f'tapu_analiz_{datetime.now().strftime("%Y%m%d")}.log')

# Logging konfigürasyonu
logger = logging.getLogger()
logger.setLevel(logging.INFO)  # DEBUG yerine INFO kullanarak log miktarını azaltıyoruz

# Rotating file handler oluştur
# maxBytes: maksimum dosya boyutu (örn: 5MB)
# backupCount: kaç yedek dosya tutulacak
file_handler = RotatingFileHandler(
    log_filename,
    maxBytes=5*1024*1024,  # 5MB
    backupCount=3,
    encoding='utf-8'
)

# Format belirle
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
file_handler.setFormatter(formatter)

# Konsol handler'ı
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)

# Handler'ları ekle
logger.addHandler(file_handler)
logger.addHandler(console_handler)

# Logging başlangıç mesajı
logging.info("=== Uygulama başlatıldı ===")

class TesisProcessor:
    def __init__(self, db_path):
        self.db_path = db_path

    def process_all(self):
        """Tüm tesis tarih işlemlerini yürütür"""
        try:
            if not self.duzenle_tesis_tarih():
                return False
            if not self.duzelt_ve_kontrol_tesis_tarih():
                return False
            if not self.guncelle_sn_ve_malik_bilgileri():
                return False
            if not self.guncelle_bos_rehin_bilgileri():
                return False
            if not self.yevmiye_guncelle():
                return False
            return True
        except Exception as e:
            logging.error(f"Tesis işleme hatası: {str(e)}")
            return False

    def duzenle_tesis_tarih(self):
        """İpotek verilerindeki tesis tarih bilgilerini düzenler"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
                                                                                                   
        try:
            kaldirilacak_ifadeler = [
               'TDesis Tarih - Yev ',
               'Tesis Tarih - Yev ',
               'Tescil Tarih - Yev Terkin',
               'Tescil Tarih - Yev ',
               'Tesis Kurum Tarih- Yevmiye',
               'Tesis Tarih - Yev ',
               'Tarih- Yevmiye',
               'Tarih - Yevmiye',
               'Tesis Kurum',
               'Tesis Tarih',
               'Tescil',
               'Tarih -',               
               'Yev Terkin',
               'Yev',
               'Terkin',
               'Tarih - Yev Terkin',
               '- Yev Terkin',
            ]
        
            for ifade in kaldirilacak_ifadeler:
                cursor.execute("""
                    UPDATE ipotek_verileri 
                    SET tesis_tarih = REPLACE(tesis_tarih, ?, ''),
                        tescil_tarih = REPLACE(tescil_tarih, ?, '')
                    WHERE tesis_tarih LIKE ? OR tescil_tarih LIKE ?
                """, (ifade, ifade, f'%{ifade}%', f'%{ifade}%'))

            # Doğru format kontrolü için regex pattern
            format_pattern = r'^[A-Za-zğüşıöçĞÜŞİÖÇ\s]+\([A-Za-zğüşıöçĞÜŞİÖÇ\s]+\)\s*-\s*\d{2}-\d{2}-\d{4}\s+\d{2}:\d{2}\s*-\s*\d+'

            # Tüm kayıtları al
            cursor.execute("SELECT rowid, tesis_tarih, tescil_tarih FROM ipotek_verileri")
            kayitlar = cursor.fetchall()

            for rowid, tesis_tarih, tescil_tarih in kayitlar:
                tesis_format_uygun = bool(re.match(format_pattern, str(tesis_tarih).strip()))
                tescil_format_uygun = bool(re.match(format_pattern, str(tescil_tarih).strip()))

                # Her iki sütunda da veri varsa format kontrolü yap
                if tesis_tarih and tescil_tarih:
                    if tescil_format_uygun and not tesis_format_uygun:
                        # Tescil tarihi doğru formatta, tesis tarihi yanlış
                        cursor.execute("""
                            UPDATE ipotek_verileri 
                            SET tesis_tarih = tescil_tarih
                            WHERE rowid = ?
                        """, (rowid,))
                    elif tesis_format_uygun and not tescil_format_uygun:
                        # Tesis tarihi doğru formatta, tescil tarihi yanlış
                        cursor.execute("""
                            UPDATE ipotek_verileri 
                            SET tescil_tarih = tesis_tarih
                            WHERE rowid = ?
                        """, (rowid,))
                # Sadece birinde veri varsa ve format uygunsa diğerine kopyala
                elif tescil_tarih and tescil_format_uygun and not tesis_tarih:
                    cursor.execute("""
                        UPDATE ipotek_verileri 
                        SET tesis_tarih = tescil_tarih
                        WHERE rowid = ?
                    """, (rowid,))
                elif tesis_tarih and tesis_format_uygun and not tescil_tarih:
                    cursor.execute("""
                        UPDATE ipotek_verileri 
                        SET tescil_tarih = tesis_tarih
                        WHERE rowid = ?
                    """, (rowid,))

            conn.commit()
            return True

        except Exception as e:
            logging.error(f"Tarih düzenleme hatası: {str(e)}")
            if conn:
                conn.rollback()
            return False

        finally:
            if conn:
                conn.close()

    def duzelt_ve_kontrol_tesis_tarih(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
           # Tescil tarih temizleme işlemi
           #print("\nTescil tarih temizleme başlıyor...")
           temizlenecek_ifadeler = [
               'Tescil Tarih - Yev Terkin',
               'Tescil Tarih - Yev ',
               'Tesis Kurum Tarih- Yevmiye',
               'Tesis Tarih - Yev ',
               'Tarih- Yevmiye',
               'Tarih - Yevmiye',
               'Tesis Kurum',
               'Tesis Tarih',
               'Tescil',
               'Tarih -',
               ' - ',
               'Yev Terkin',
               'Yev',
               'Terkin',
               'Tarih - Yev Terkin',
               '- Yev Terkin',
           ]
       
           for ifade in temizlenecek_ifadeler:
               cursor.execute("""
                   UPDATE ipotek_verileri 
                   SET tescil_tarih = REPLACE(tesis_tarih, ?, '')
                   WHERE tescil_tarih LIKE ?
               """, (ifade, f'%{ifade}%'))

           for ifade in temizlenecek_ifadeler:
               cursor.execute("""
                   UPDATE ipotek_verileri 
                   SET tescil_tarih = REPLACE(tescil_tarih, ?, '')
                   WHERE tescil_tarih LIKE ?
               """, (ifade, f'%{ifade}%'))


               #print(f"'{ifade}' ifadesi temizlendi")

           format_pattern = r'^.+ - \d{2}-\d{2}-\d{4} \d{2}:\d{2} - \d+$'
       
           # Önce hatalı kayıtları bul ve tescil_tarih kontrolü yap
           cursor.execute("""
               SELECT rowid, tesis_tarih, tescil_tarih 
               FROM ipotek_verileri 
               WHERE tesis_tarih IS NOT NULL 
               AND tesis_tarih != ''
           """)
       
           guncelleme_listesi = []
           hatali_kayitlar = []
       
           #print("\nKayıtlar kontrol ediliyor...")
           for rowid, tesis_tarih, tescil_tarih in cursor.fetchall():
               tesis_uygun = bool(re.match(format_pattern, str(tesis_tarih).strip()))
               tescil_uygun = bool(re.match(format_pattern, str(tescil_tarih).strip()))
           
               if not tesis_uygun and tescil_uygun:
                   guncelleme_listesi.append((rowid, tescil_tarih))
                   """
                   print(f"\nGüncellenecek kayıt bulundu:")
                   print(f"ID: {rowid}")
                   print(f"Eski tesis_tarih: {tesis_tarih}")
                   print(f"Yeni tesis_tarih olacak: {tescil_tarih}")
                   """
               elif not tesis_uygun:
                   hatali_kayitlar.append((rowid, tesis_tarih, tescil_tarih))

           # Güncelleme işlemini yap
           if guncelleme_listesi:
               #print(f"\nToplam {len(guncelleme_listesi)} kayıt güncellenecek...")
               for rowid, yeni_deger in guncelleme_listesi:
                   cursor.execute("""
                       UPDATE ipotek_verileri 
                       SET tesis_tarih = ? 
                       WHERE rowid = ?
                   """, (yeni_deger, rowid))
               #print("Güncelleme tamamlandı.")
           """
           # Hala hatalı olan kayıtları göster
           if hatali_kayitlar:
               print("\nHala hatalı formatta olan ve düzeltilemeyecek kayıtlar:")
               print("ID\tTesis Tarih || Tescil Tarih")
               print("-" * 80)
               for rowid, tesis_tarih, tescil_tarih in hatali_kayitlar:
                   print(f"{rowid}\t{tesis_tarih} || {tescil_tarih}")
               print(f"\nToplam {len(hatali_kayitlar)} adet hatalı kayıt kaldı.")
           else:
               print("\nTüm kayıtlar düzeltildi veya uygun formatta.")

           conn.commit()
           """
           return True

        except Exception as e:
           #print(f"Hata oluştu: {str(e)}")
           conn.rollback()
           return False
   
        finally:
           conn.close()

    def guncelle_sn_ve_malik_bilgileri(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
    
        try:
            # 1. Boş SN bilgilerini doldur
            cursor.execute("""
                SELECT rowid, hisse_pay_payda 
                FROM ipotek_verileri 
                WHERE (sn_bilgisi IS NULL OR sn_bilgisi = '')
                AND hisse_pay_payda LIKE '%(SN:%'
            """)
        
            # SN numaralarını çıkar ve güncelle
            sn_pattern = r'\(SN:(\d+)\)'
            for rowid, hisse_pay_payda in cursor.fetchall():
                match = re.search(sn_pattern, hisse_pay_payda)
                if match:
                    sn_no = match.group(1)
                    cursor.execute("""
                        UPDATE ipotek_verileri 
                        SET sn_bilgisi = ? 
                        WHERE rowid = ?
                    """, (sn_no, rowid))
        
            # 2. Eşleştirme ve güncelleme için kayıtları al
            cursor.execute("""
                SELECT rowid, tasinmaz_kimlik, sn_bilgisi 
                FROM ipotek_verileri 
                WHERE sn_bilgisi IS NOT NULL 
                AND sn_bilgisi != ''
            """)
        
            for rowid, tasinmaz_kimlik, sn_bilgisi in cursor.fetchall():
                # Eşleşen tapu kaydını bul
                cursor.execute("""
                    SELECT hucreno_2, hucreno_4 
                    FROM tapu_verileri 
                    WHERE tasinmaz_kimlik = ?
                    AND hucreno_2 LIKE ?
                """, (tasinmaz_kimlik, f'%(SN:{sn_bilgisi})%'))
            
                tapu_kayit = cursor.fetchone()
                if tapu_kayit:
                    hucreno_2, hucreno_4 = tapu_kayit
                
                    # SN bilgisini temizle
                    temiz_borclu = re.sub(r'\(SN:\d+\)\s*', '', hucreno_2)
                
                    # Güncelleme yap
                    cursor.execute("""
                        UPDATE ipotek_verileri 
                        SET borclu_malik = ?,
                            hisse_pay_payda = ?
                        WHERE rowid = ?
                    """, (temiz_borclu, hucreno_4, rowid))
        
            conn.commit()
            return True

        except Exception as e:
            #print(f"Hata oluştu: {str(e)}")
            conn.rollback()
            return False
    
        finally:
            conn.close()

    def guncelle_bos_rehin_bilgileri(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
    
        try:
            #print("\nBOŞ REHİN GÜNCELLEME İŞLEMİ BAŞLIYOR...")
        
            # 1. BosRehin olan ve diğer alanları boş olan kayıtları bul
            cursor.execute("""
                SELECT sayfano, satirno FROM tapu_verileri 
                WHERE hucreno_1 = 'BosRehin' 
                AND (hucreno_2 IS NULL OR hucreno_2 = '' OR hucreno_2 = ' ')
                AND (hucreno_3 IS NULL OR hucreno_3 = '' OR hucreno_3 = ' ')
                AND (hucreno_4 IS NULL OR hucreno_4 = '' OR hucreno_4 = ' ')
                AND (hucreno_5 IS NULL OR hucreno_5 = '' OR hucreno_5 = ' ')
                AND (hucreno_6 IS NULL OR hucreno_6 = '' OR hucreno_6 = ' ')
            """)
        
            bos_rehin_kayitlari = cursor.fetchall()
            #print(f"Bulunan boş rehin kayıt sayısı: {len(bos_rehin_kayitlari)}")
        
            guncellenen_kayit_sayisi = 0
        
            for sayfano, satirno in bos_rehin_kayitlari:
                #print(f"\nİŞLENEN KAYIT -> Sayfa: {sayfano}, Satır: {satirno}")
            
                # 2. İki satır sonrasındaki veriyi al
                hedef_satirno = satirno + 2
            
                cursor.execute("""
                    SELECT hucreno_1 FROM tapu_verileri 
                    WHERE sayfano = ? AND satirno = ?
                """, (sayfano, hedef_satirno))
            
                sonuc = cursor.fetchone()
                if not sonuc or not sonuc[0]:
                    #print(f"  - 2 satır sonrası için veri bulunamadı")
                    continue
            
                uzun_metin = sonuc[0]
                #print(f"  - Bulunan metin: {uzun_metin[:100]}...")
            
                # SN numarasını bul
                sn_match = re.search(r'SN:(\d+)', uzun_metin)
                if not sn_match:
                    #print("  - SN numarası bulunamadı")
                    continue
            
                sn_numarasi = sn_match.group(1)
                #print(f"  - Çıkarılan SN numarası: {sn_numarasi}")
            
                # 3. İpotek verilerini güncelle
                cursor.execute("""
                    UPDATE ipotek_verileri 
                    SET alacakli = 'Boş Rehin Bilgisi'
                    WHERE sn_bilgisi = ? 
                    AND LOWER(alacakli) = LOWER('İpoteğin Konulduğu Hisse Bilgisi')
                """, (sn_numarasi,))
            
                etkilenen_kayit = cursor.rowcount
                if etkilenen_kayit > 0:
                    guncellenen_kayit_sayisi += etkilenen_kayit
                    #print(f"  - Güncellenen kayıt sayısı: {etkilenen_kayit}")
                #else:
                    #print("  - Güncelleme kriterlerine uygun kayıt bulunamadı")
        
            conn.commit()
            #print(f"\nİŞLEM TAMAMLANDI")
            #print(f"Toplam güncellenen kayıt sayısı: {guncellenen_kayit_sayisi}")
            return True
        
        except Exception as e:
            #print(f"\nHATA OLUŞTU: {str(e)}")
            conn.rollback()
            return False
        
        finally:
            conn.close()

    def yevmiye_guncelle(self):        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            def temizle_filigran(text):
                if not text:
                    return ""
            
                # Filigran harfleri
                filigran_harfler = ["B", "İ", "L", "G", "İ", "A", "M", "A", "Ç", "L", "I", "D", "R"]
            
                # Sayılar arasındaki filigranları temizle
                temiz_metin = ""
                i = 0
                while i < len(text):
                    if (i > 0 and i < len(text) - 1 and  # Önceki ve sonraki karakter var mı
                        text[i] in filigran_harfler and  # Şu anki karakter filigran mı
                        (text[i-1].isdigit() or text[i-1] in ':./-') and  # Önceki karakter sayı veya ayraç mı
                        (text[i+1].isdigit() or text[i+1] in ':./-')):    # Sonraki karakter sayı veya ayraç mı
                        i += 1  # Filigran karakterini atla
                    else:
                        temiz_metin += text[i]
                        i += 1

                # Tam filigran kelimelerini temizle
                filigran_kelimeler = ["BİLGİ", "AMAÇLIDIR", "BİLGİ AMAÇLIDIR"]
                for kelime in filigran_kelimeler:
                    temiz_metin = temiz_metin.replace(kelime, "")

                return temiz_metin.strip()

            pattern = r"""
                # İlk kısım (yer bilgisi) - zorunlu değil
                (?:[^-\d]+)?
            
                # Tarih kısmı (zorunlu)
                (\d{2}[-\./]\d{2}[-\./]\d{4})
            
                # Saat kısmı (opsiyonel)
                (?:\s+(\d{1,2}:\d{1,2}))?
            
                # Son kısım - yevmiye no
                .*?[-\s]+(\d+)\s*$
            """
            cursor.execute(f"""
                SELECT rowid, hucreno_3, hucreno_4, 
                       hucreno_5, hucreno_6, hucreno_7, hucreno_8, hucreno_9
                FROM tapu_verileri 
                WHERE hucreno_1 IS NOT NULL
                AND (takyidat_1 IS NULL OR takyidat_1 = '')
                AND (baslikontrol IS NOT 'EVET')
            """)
        
            kayitlar = cursor.fetchall()
            guncellenen = 0
        
            for kayit in kayitlar:
                kayit_id = kayit[0]
            
                for hucre_deger in kayit[1:]:
                    if not hucre_deger:
                        continue
                
                    # Önce filigranları temizle
                    temiz_deger = temizle_filigran(str(hucre_deger))
                
                    # Parantezleri ve içindekileri temizle
                    temiz_deger = re.sub(r'\([^)]*\)', '', temiz_deger)
                
                    # Regex ile eşleşme kontrolü
                    match = re.search(pattern, temiz_deger, re.VERBOSE)
                
                    if match:
                        try:
                            # İşlem yerini al - ilk tarihten önceki kısım
                            islem_yeri = temiz_deger.split(match.group(1))[0].strip()
                            islem_yeri = re.sub(r'[-\s]+$', '', islem_yeri)  # Sondaki tire ve boşlukları temizle
                        
                            # Tarih ve saat formatını ayarla
                            tarih = match.group(1)  # Tarih kısmı
                            saat = match.group(2)   # Saat kısmı (opsiyonel)
                        
                            # Tarihi noktalı formata çevir
                            tarih_parcalar = re.split('[-./]', tarih)
                            formatli_tarih = f"{tarih_parcalar[0]}.{tarih_parcalar[1]}.{tarih_parcalar[2]}"
                        
                            # Saat varsa ekle, yoksa sadece tarih
                            tam_tarih = f"{formatli_tarih} {saat}" if saat else formatli_tarih
                        
                            # Yevmiye numarası artık group(3)'te
                            yevmiye_no = match.group(3)

                            # Verileri güncelle
                            cursor.execute("""
                                UPDATE tapu_verileri 
                                SET takyidat_1 = ?,  -- Yevmiye No
                                    takyidat_2 = ?,  -- Tarih (ve varsa Saat)
                                    takyidat_3 = ?   -- İşlem Yeri
                                WHERE rowid = ?
                            """, (yevmiye_no, tam_tarih, islem_yeri, kayit_id))
                        
                            guncellenen += 1
                            break
                        except Exception as e:
                            print(f"Veri işleme hatası: {str(e)}")
                            continue

            conn.commit()
            return f"İşlem başarıyla tamamlandı. {guncellenen} kayıt güncellendi."
        
        except Exception as e:
            print(f"Genel hata: {str(e)}")
            return f"Genel hata oluştu: {str(e)}"
    
        finally:
            if conn:
                conn.close()

class TableAnalyzer:
    def __init__(self):
        self.common_headers = [
            "TAPU KAYIT BİLGİSİ",
            "MUHDESAT BİLGİLERİ",
            "TAŞINMAZA AİT ŞERH BEYAN İRTİFAK BİLGİLERİ",
            "EKLENTİ BİLGİLERİ",
            "MÜLKİYET BİLGİLERİ",
            "TEFERRUAT BİLGİLERİ",
            "MÜLKİYETE AİT ŞERH BEYAN İRTİFAK BİLGİLERİ",
            "MÜLKİYETE AİT REHİN BİLGİLERİ"
        ]
        self.margin_threshold = 50  # Sol kenar boşluğu

    def tapu_belgesi_kontrol(self, text):
        """
        Belgenin tapu belgesi olup olmadığını kontrol eder
        """
        # Basit başlık kontrolü
        if "TAPU KAYIT BİLGİSİ" not in text:
            logging.error("TAPU KAYIT BİLGİSİ başlığı bulunamadı")
            return False

        # Zorunlu alanların kontrolü
        zorunlu_alanlar = [
            "Zemin Tipi:",
            "Taşınmaz Kimlik No:",
            "İl/İlçe:",
            "Ada/Parsel:"
        ]

        # En az 3 zorunlu alan bulunması yeterli
        bulunan_alan_sayisi = sum(1 for alan in zorunlu_alanlar if alan in text)
        if bulunan_alan_sayisi >= 3:
            logging.info("Yeni Takbis ")
            return True

        logging.error("Yetersiz zorunlu alan sayısı")
        return False

    def tasinmaz_kimlik_no_al(self, text):
        """
        Metin içinden taşınmaz kimlik numarasını çıkarır
        """
        try:
            # Direkt olarak "Taşınmaz Kimlik No:" satırını ara
            for line in text.split('\n'):
                if "Taşınmaz Kimlik No:" in line:
                    # Numarayı AT'den önceki kısımdan al
                    kimlik_no = line.split('Taşınmaz Kimlik No:')[1].split('AT')[0].strip()
                    logging.info(f"Taşınmaz Kimlik No bulundu: {kimlik_no}")
                    return kimlik_no

            logging.error("Taşınmaz Kimlik No satırı bulunamadı")
            return None

        except Exception as e:
            logging.error(f"Taşınmaz Kimlik No çıkarma hatası: {str(e)}")
            return None

    def create_ext_koordinat_table(self, db_path):
        """Genişletilmiş koordinat tablosunu oluştur"""
        try:
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
    
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS koordinat_bilgileri_ext (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tapu_rowid INTEGER,
                    hucreno_1_deger TEXT,
                    y_koordinat REAL,
                    tasinmaz_kimlik TEXT,
                    sayfano INTEGER,
                    satirno INTEGER,
                    baslik_deger INTEGER
                )
            """)
    
            conn.commit()
            return True
        except Exception as e:
            logging.error(f"Ext koordinat tablosu oluşturma hatası: {str(e)}")
            return False
        finally:
            if conn:
                conn.close()

    def extract_fitz_coordinates(self, pdf_path, tasinmaz_kimlik):
        """Fitz ile koordinatları çıkar ve kaydet"""
        try:
            doc = fitz.open(pdf_path)
            conn = sqlite3.connect("veritabani.db")
            cursor = conn.cursor()
        
            # Atlanacak ifadeler listesi
            skip_texts = [
                "veya Web Tapu anasayfasından",
                "veya Web Tapu",
                "kodunu Online İşlemler",
                "kodunu Online",
                "https://webtapu.tkgm.gov.tr",
                "alanına yazarak doğrulayabilirsiniz",
                "adresinden)",
                "Kaydı Oluşturan:",
                "Bu belgeyi akıllı telefonunuzdan karekod tarama programları ile aşağıdaki barkodu",
                "taratarak;",
                "akıllı telefonunuzdan",
                "BU BELGE TOPLAM",
                "BİLGİ AMAÇLIDIR",
                "İl/İlçe:",
                "Tesis Tarih - Yev",
                "Tesis Kurum Tarih-Yevmiye",
                "Tesis Kurum Tarih-",
                "Tescil Tarih - Yev",
                "Terkin Sebebi-",
                "-",
                "Ş/B/İ",
                "_",
                "Zemin Tipi:",
                "Başvuru No",
                "Bağımsız Bölüm Nitelik:",
                "Bağımsız Bölüm Net",
                "Bağımsız Bölüm Brüt",
                "Arsa Pay/Payda:"
                ]


            satirno = 0
            for page_num, page in enumerate(doc, 1):
                blocks = page.get_text("dict")["blocks"]
            
                for block in blocks:
                    if "lines" in block:
                        for line in block["lines"]:
                            for span in line["spans"]:
                                text = span["text"].strip()
                                if text:  # Boş olmayan text
                                    # Atlanacak ifadeleri kontrol et
                                    skip = False
                                    for skip_text in skip_texts:
                                        if skip_text.lower() in text.lower():
                                            skip = True
                                            break
                                
                                    if not skip:
                                        y_coord = span["origin"][1]  # y koordinatı
                                        satirno += 1
                                    
                                        cursor.execute("""
                                            INSERT INTO koordinat_bilgileri_ext 
                                            (tapu_rowid, hucreno_1_deger, y_koordinat, tasinmaz_kimlik, 
                                             sayfano, satirno, baslik_deger)
                                            VALUES (?, ?, ?, ?, ?, ?, ?)
                                        """, (None, text, y_coord, tasinmaz_kimlik, page_num, satirno, None))
        
            conn.commit()
            return True
        
        except Exception as e:
            logging.error(f"Fitz koordinat çıkarma hatası: {str(e)}")
            return False
        finally:
            if 'doc' in locals():
                doc.close()
            if 'conn' in locals():
                conn.close()

    def tasinmaz_kayit_kontrol(self, tasinmaz_no, db_yolu):
        """
        Taşınmaz numarasının veritabanında olup olmadığını kontrol eder
        """
        import sqlite3
        
        try:
            conn = sqlite3.connect(db_yolu)
            cursor = conn.cursor()
            
            # Veritabanında taşınmazı ara
            cursor.execute("""
                SELECT COUNT(*) FROM tapu_verileri 
                WHERE tasinmaz_kimlik = ?
            """, (tasinmaz_no,))
            
            sayi = cursor.fetchone()[0]
            return sayi > 0
            
        except Exception as e:
            #print(f"Veritabanı kontrol hatası: {str(e)}")
            return False
            
        finally:
            if conn:
                conn.close()

    def tapu_belgesi_dogrula(self, text, db_yolu):
        """
        Ana doğrulama fonksiyonu - hem belge türünü hem mükerrer kaydı kontrol eder
        """
        # Önce tapu belgesi mi kontrol et
        if not self.tapu_belgesi_kontrol(text):
            return False, "Bu belge geçerli bir tapu belgesi değil."
        
        # Taşınmaz kimlik no'yu bul
        tasinmaz_no = self.tasinmaz_kimlik_no_al(text)
        if not tasinmaz_no:
            return False, "Taşınmaz Kimlik No bulunamadı."
        
        # Mükerrer kayıt kontrolü
        if self.tasinmaz_kayit_kontrol(tasinmaz_no, db_yolu):
            return False, f"Bu taşınmaz (Kimlik No: {tasinmaz_no}) zaten veritabanında mevcut."
        
        return True, tasinmaz_no
  
    def clean_filigran(self, text):
        """Filigran harflerini temizler ve metni düzenler"""
        if not text:
            return ""
            
        filigran_harfler = ["B", "İ", "L", "G", "İ", "A", "M", "A", "Ç", "L", "I", "D", "R"]
        
        # Tek başına duran filigran harflerini temizle
        cleaned_text = ""
        words = text.split()
        for word in words:
            if word not in filigran_harfler:
                cleaned_text += word + " "
                
        # Fazla boşlukları temizle
        cleaned_text = " ".join(cleaned_text.split())
        return cleaned_text

    def validate_header_position(self, header, table_bbox, text_line):
        """
        Başlığın konumunun geçerli olup olmadığını kontrol eder
    
        Args:
            header (str): Kontrol edilecek başlık metni
            table_bbox (tuple): Tablo sınırlarını içeren koordinatlar
            text_line (dict): Başlık satırının koordinat ve özelliklerini içeren sözlük
        
        Returns:
            bool: Başlığın konumu geçerliyse True, değilse False
        """
        # Boş veya geçersiz başlık kontrolü
        if not header or not isinstance(header, str):
            return False
        
        # TAPU KAYIT BİLGİSİ için özel durum
        if header == "TAPU KAYIT BİLGİSİ":
            return text_line['x0'] <= 50  # Sol kenar kontrolü
        
        # MÜLKİYETE AİT REHİN BİLGİLERİ için özel durum
        if header == "MÜLKİYETE AİT REHİN BİLGİLERİ":
            # Sayfa sonunda olabilir, bu yüzden daha esnek kontrol
            return True
        
        # Başlık tablonun üstünde olmalı ve sol kenara yakın olmalı
        if text_line['top'] >= table_bbox[1] or text_line['x0'] > 50:
            return False
        
        return True

    def detect_header(self, page, table_bbox):
        """Tablonun üstündeki başlığı tespit eder"""
        try:
            texts = page.extract_text_lines()
            header = None
            header_y = float('inf')
            margin_threshold = 50  # Sol kenar boşluğu
            prev_header = None  # Önceki sayfadan kalan başlık kontrolü için
        
            # Başlık grupları
            no_underline_headers = [
                "TAPU KAYIT BİLGİSİ"  # Bu başlık her zaman ilk sayfada ve altında çizgi yok
            ]
        
            special_headers = [
                "MÜLKİYET BİLGİLERİ",
                "MÜLKİYETE AİT ŞERH BEYAN İRTİFAK BİLGİLERİ",
                "MÜLKİYETE AİT REHİN BİLGİLERİ",
                "TAŞINMAZA AİT ŞERH BEYAN İRTİFAK BİLGİLERİ"
            ]
        
            secondary_headers = [
                "TEFERRUAT BİLGİLERİ",
                "EKLENTİ BİLGİLERİ",
                "MUHDESAT BİLGİLERİ"
            ]
        
            for text_line in texts:
                raw_text = text_line['text'].strip()
                text = self.clean_filigran(raw_text)
            
                if not text:
                    continue
            
                # TAPU KAYIT BİLGİSİ özel kontrolü
                if text in no_underline_headers:
                    if text_line['x0'] <= margin_threshold:
                        return text
            
                # Başlık kontrolü
                if self.is_potential_header(text):
                    if self.validate_header_position(text, table_bbox, text_line):
                        # Yatay çizgi kontrolü
                        has_underline = False
                        for line in page.lines:
                            if (line['top'] > text_line['bottom'] and 
                                line['top'] < text_line['bottom'] + 30 and
                                line['x0'] <= margin_threshold + 10):
                                if abs(line['y0'] - line['y1']) < 5:  # Yatay çizgi
                                    has_underline = True
                                    break
                    
                        # Özel başlıklar veya altı çizili başlıklar için
                        if text in special_headers:  # Ana başlıklar her durumda kabul edilir
                            if text_line['top'] < header_y:
                                prev_header = header
                                header = text
                                header_y = text_line['top']
                        elif text in secondary_headers and has_underline:  # İkincil başlıklar çizgi gerektirir
                            if text_line['top'] < header_y:
                                prev_header = header
                                header = text
                                header_y = text_line['top']
                
            return header if header else prev_header
            
        except Exception as e:
            logging.error(f"Başlık tespiti hatası: {str(e)}")
            return None

    def is_potential_header(self, text):
        """Metnin başlık olma potansiyelini kontrol eder"""
        if not text or len(text) < 5:
            return False
            
        # Kesin başlıklar
        exact_headers = [
            "TAPU KAYIT BİLGİSİ",
            "MÜLKİYETE AİT REHİN BİLGİLERİ",
            "TAŞINMAZA AİT ŞERH BEYAN İRTİFAK BİLGİLERİ",
            "MÜLKİYET BİLGİLERİ",
            "MÜLKİYETE AİT ŞERH BEYAN İRTİFAK BİLGİLERİ",
            "EKLENTİ BİLGİLERİ",
            "MUHDESAT BİLGİLERİ"
        ]
        
        # Tam eşleşme kontrolü
        if text in exact_headers:
            return True
            
        # Diğer başlık kontrolleri için mevcut mantığı koru
        if text in self.common_headers:
            return True
            
        if text.isupper() and ("BİLGİ" in text or "KAYIT" in text):
            return True
            
        return False

    def process_table(self, table, page):
        processed_rows = []
        try:
            table_data = table.extract()
            table_cells = table.cells
    
            if not table_data or not table_cells:
                return []

            # Sayfanın yüksekliğini al
            page_height = page.height
        
            # Sayfadaki kelimeleri al
            words = page.extract_words(
                x_tolerance=3, 
                y_tolerance=3,
                keep_blank_chars=True,
                use_text_flow=False
            )
        
            for row_idx, row_data in enumerate(table_data):
                processed_cells = []
            
                # İlk hücre içeriği ve y koordinatını bul 
                first_cell = str(row_data[0]).strip() if row_data else ''
                y_coord = None
            
                if first_cell:
                    # Y koordinatını bulmak için alternatif yöntemler
                    # 1. Kelimelerden bulma
                    matching_words = [w for w in words if first_cell in w['text']]
                    if matching_words:
                        y_coord = matching_words[0]['top']
                
                    # 2. Tablo hücresinden alma (backup)
                    if y_coord is None and table_cells and len(table_cells) > row_idx:
                        cell = table_cells[row_idx][0]
                        if cell and hasattr(cell, 'bbox'):
                            y_coord = page_height - cell.bbox[3]  # Sayfanın altından mesafe
                
                    # 3. Son çare: Önceki satırdan tahmin
                    if y_coord is None and processed_rows:
                        last_y = processed_rows[-1]['y_position']
                        if last_y is not None:
                            y_coord = last_y + 20  # Ortalama satır yüksekliği
            
                # Hücreleri işle
                for col_idx, cell_content in enumerate(row_data):
                    try:
                        cell = table_cells[row_idx][col_idx] if row_idx < len(table_cells) and col_idx < len(table_cells[row_idx]) else None
                    
                        content = str(cell_content).strip() if cell_content is not None else ''
                        content = self.analyze_and_clean_filigran(content)
                        content = content.replace('BİLGİ AMAÇLIDIR', '').strip()
                        content = ' '.join(content.split())
                    
                        processed_cells.append({
                            'content': content,
                            'bbox': cell.bbox if cell and hasattr(cell, 'bbox') else None
                        })
                    
                    except Exception as cell_error:
                        logging.error(f"Hücre işleme hatası [{row_idx}][{col_idx}]: {str(cell_error)}")
                        processed_cells.append({
                            'content': '',
                            'bbox': None
                        })
            
                if any(cell['content'] for cell in processed_cells):
                    processed_rows.append({
                        'cells': processed_cells,
                        'y_position': y_coord
                    })
        
            return processed_rows
        
        except Exception as e:
            logging.error(f"Tablo işleme hatası: {str(e)}")
            return []

    def analyze_and_clean_filigran(self, text):
        """Filigran harflerini ve kelimelerini temizler"""
        if not text:
            return ""
        
        # Filigran harfleri
        filigran_harfler = ["B", "İ", "L", "G", "İ", "A", "M", "A", "Ç", "L", "I", "D", "R"]
    
        # Tam filigran kelimeleri
        filigran_kelimeler = [
            "BİLGİ AMAÇLIDIR",
            "BİLGİ  AMAÇLIDIR",  # Çift boşluklu versiyon
            "BİLGİAMAÇLIDIR",    # Boşluksuz versiyon
        ]
    
        # Önce tam kelimeleri temizle
        for kelime in filigran_kelimeler:
            text = text.replace(kelime, '')
    
        # Kelimeleri parçala
        words = text.split()
        cleaned_words = []
    
        for word in words:
            # Eğer kelime tek bir filigran harfi değilse ekle
            if word not in filigran_harfler and not (len(word) == 1 and word.upper() in filigran_harfler):
                cleaned_words.append(word)
    
        # Temizlenmiş metni birleştir
        cleaned_text = ' '.join(cleaned_words)
    
        # Fazla boşlukları temizle
        cleaned_text = re.sub(r'\s+', ' ', cleaned_text)
    
        return cleaned_text.strip()

def enhanced_analyze_pdf(pdf_path: str, db_path: str, tasinmaz_kimlik_no: str) -> bool:
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        table_analyzer = TableAnalyzer()

        # Önce mevcut kolonları kontrol et
        cursor.execute("PRAGMA table_info(tapu_verileri)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns]

        # y_koordinat ve baslik_deger kolonları yoksa ekle
        if 'y_koordinat' not in column_names:
            cursor.execute("ALTER TABLE tapu_verileri ADD COLUMN y_koordinat REAL")
        if 'baslik_deger' not in column_names:
            cursor.execute("ALTER TABLE tapu_verileri ADD COLUMN baslik_deger INTEGER")

        # Tapu verileri tablosunu oluştur
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS tapu_verileri (
                sayfano INTEGER NOT NULL,
                satirno INTEGER NOT NULL,
                hucreno_1 TEXT,
                hucreno_2 TEXT,
                hucreno_3 TEXT,
                hucreno_4 TEXT,
                hucreno_5 TEXT,
                hucreno_6 TEXT,
                hucreno_7 TEXT,
                hucreno_8 TEXT,
                hucreno_9 TEXT,
                hucreno_10 TEXT,
                hucreno_11 TEXT,
                baslikontrol TEXT DEFAULT 'HAYIR',
                tasinmaz_kimlik TEXT,
                takyidat_baslik TEXT,
                takyidat_1 TEXT,
                takyidat_2 TEXT,
                takyidat_3 TEXT,
                y_koordinat REAL,
                baslik_deger INTEGER
            )
        """)

        # Koordinat tablosunu oluştur
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS koordinat_bilgileri_ext (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tapu_rowid INTEGER,
                hucreno_1_deger TEXT,
                y_koordinat REAL,
                tasinmaz_kimlik TEXT,
                sayfano INTEGER,
                satirno INTEGER
            )
        """)

        current_row = 0
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                tables = page.find_tables(
                    table_settings={
                        "vertical_strategy": "lines",
                        "horizontal_strategy": "lines",
                        "intersection_y_tolerance": 3,
                        "intersection_x_tolerance": 3
                    }
                )
                
                for table in tables:
                    processed_rows = table_analyzer.process_table(table, page)
                
                    for row_data in processed_rows:
                        values = [page_num, current_row]
                        cell_values = []
                    
                        # Dict'ten sadece content değerlerini al
                        for cell in row_data['cells']:
                            cell_values.append(cell['content'])
                    
                        # 11 hücreye tamamla
                        while len(cell_values) < 11:
                            cell_values.append('')
                        
                        values.extend(cell_values)
                        values.extend([
                            "HAYIR",
                            tasinmaz_kimlik_no,
                            "",  # takyidat_baslik
                            "", "", ""
                        ])
                        
                        # Y koordinat bilgisini ekle
                        if 'y_position' in row_data:
                            values.extend([row_data['y_position'], None])  # y_koordinat ve boş baslik_deger
                        else:
                            values.extend([None, None])
                    
                        # Tapu verilerini kaydet
                        placeholders = ','.join(['?' for _ in range(len(values))])
                        cursor.execute(f"""
                            INSERT INTO tapu_verileri VALUES ({placeholders})
                        """, values)
                        tapu_rowid = cursor.lastrowid                  
                        current_row += 1
                                           
        conn.commit()
        return True
        
    except Exception as e:
        logging.error(f"PDF analiz hatası: {str(e)}")
        return False
        
    finally:
        if conn:
            conn.close()

class ProcessingDialog(QDialog):
    def __init__(self, total_files, parent=None):
        super().__init__(parent)
        self.setWindowModality(Qt.ApplicationModal)
        self.setWindowFlags(Qt.Window | Qt.WindowCloseButtonHint)
        self.setWindowTitle('Toplu İşlem')
        self.setMinimumWidth(400)
        
        # İşlem durumu flag'i
        self.is_processing = False
        
        # Layout oluştur
        layout = QVBoxLayout()
        
        self.file_count_label = QLabel(f"Toplam Dosya: {total_files}")
        self.current_file_label = QLabel("İşlenen Dosya: ")
        self.progress = QProgressBar()
        self.progress.setMaximum(total_files)
        
        self.details = QTextEdit()
        self.details.setReadOnly(True)
        self.details.setMinimumHeight(200)
        
        layout.addWidget(self.file_count_label)
        layout.addWidget(self.current_file_label)
        layout.addWidget(self.progress)
        layout.addWidget(QLabel("İşlem Detayları:"))
        layout.addWidget(self.details)
        
        self.setLayout(layout)
                # Stil uygulaması
        self.setStyleSheet("""
            QDialog {
                background-color: #f0f0f0; /* Açık gri arka plan */
                font-family: Segoe UI, Roboto, sans-serif; /* Modern font */
                font-size: 11pt;
            }

            QLabel {
                color: #333333; /* Koyu gri metin rengi */
            }

            QProgressBar {
                border: 1px solid #cccccc;
                border-radius: 5px;
                text-align: center;
                background-color: #ffffff;
                color: #333333;
            }

            QProgressBar::chunk {
                background-color: #4CAF50; /* Yeşil ilerleme çubuğu rengi */
                border-radius: 5px;
            }

            QTextEdit {
                background-color: #ffffff;
                border: 1px solid #cccccc;
                border-radius: 5px;
                color: #333333;
            }

            QPushButton {
                background-color: #008CBA; /* Mavi buton rengi */
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
            }

            QPushButton:hover {
                background-color: #005f73; /* Daha koyu mavi buton rengi */
            }
        """)

    def closeEvent(self, event):
        if self.is_processing:
            reply = QMessageBox.question(
                self, 
                'İşlemi Sonlandır',
                'İşlem devam ediyor. Kapatmak istediğinizden emin misiniz?',
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()  # İşlem bitmişse direkt kapat

    def update_progress(self, current_file, current_count):
        self.current_file_label.setText(f"İşlenen Dosya: {os.path.basename(current_file)}")
        self.progress.setValue(current_count)
    
    def add_detail(self, message):
        self.details.append(message)

    # İşlem başlangıç ve bitiş metodları
    def start_processing(self):
        self.is_processing = True

    def finish_processing(self):
        self.is_processing = False

class TapuAnalyzerGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.table_analyzer = TableAnalyzer()
        # Log dosyası ayarlarını güncelle
        logging.basicConfig(
            filename='tapu_analiz.log',
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        # Ekrana da log göstermek için
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.DEBUG)
        logging.getLogger().addHandler(console_handler)
        
        
        self.initUI()

    def initUI(self):
        """GUI arayüzünü oluştur"""
        self.setWindowTitle('Takbis Belgeleri İşleme / Analiz / Raporlama')
        self.setGeometry(100, 100, 1200, 800)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # Menü çubuğunu oluşturma
        menubar = self.menuBar()

        # Yardım menüsü oluşturma
        help_menu = menubar.addMenu('Yardım')

        # Yardım dosyası açma eylemi
        open_help_action = QAction('Yardım Dosyası Aç', self)
        open_help_action.triggered.connect(self.open_help_file)
        help_menu.addAction(open_help_action)

        # Üst panel için ana layout
        upper_panel = QWidget()
        upper_layout = QHBoxLayout(upper_panel)  # Yatay düzen

        # Sol butonlar için dikey layout
        left_buttons = QVBoxLayout()
        self.select_multiple_btn = QPushButton('🔰 Takbisleri Seç', self)
        self.select_multiple_btn.setFixedSize(110, 30)
        self.select_multiple_btn.clicked.connect(self.select_multiple_files)
        left_buttons.addWidget(self.select_multiple_btn)

        self.batch_process_btn = QPushButton('🟥 İçeri Aktar', self)
        self.batch_process_btn.setFixedSize(110, 30)
        self.batch_process_btn.clicked.connect(self.start_batch_processing)
        left_buttons.addWidget(self.batch_process_btn)
        left_buttons.addStretch()

        # Merkez - TreeWidget
        center_layout = QVBoxLayout()
        self.tree_widget = QTreeWidget()
        self.tree_widget.setHeaderLabels(['Ada/Parsel', 'Nitelik', 'BB/Blok Bilgisi', 'İl/İlçe', 'Mahalle', 'Taşınmaz No'])
        self.tree_widget.setAlternatingRowColors(True)
        self.setup_tree_context_menu()
        self.tree_widget.itemDoubleClicked.connect(self.on_tree_double_clicked)
        center_layout.addWidget(self.tree_widget)

        # Sağ butonlar için dikey layout
        right_buttons = QVBoxLayout()
        self.multi_format_btn = QPushButton('📑 Çoklu Raporla', self)
        self.multi_format_btn.setFixedSize(110, 30)
        self.multi_format_btn.clicked.connect(self.show_multi_format)
        right_buttons.addWidget(self.multi_format_btn)

        self.analyze_takbis_btn = QPushButton('📄 Tekli Raporla', self)
        self.analyze_takbis_btn.setFixedSize(110, 30)
        self.analyze_takbis_btn.clicked.connect(self.analyze_takbis)
        right_buttons.addWidget(self.analyze_takbis_btn)

        self.clear_btn = QPushButton('🧹Temizle', self)
        self.clear_btn.setFixedSize(110, 30)
        self.clear_btn.clicked.connect(self.clear_text)
        right_buttons.addWidget(self.clear_btn)

        self.save_btn = QPushButton('💾 TXT Kaydet', self)
        self.save_btn.setFixedSize(110, 30)
        self.save_btn.clicked.connect(self.save_report)
        right_buttons.addWidget(self.save_btn)

        self.export_excel_btn = QPushButton('📊 Hisse Tablosu', self)
        self.export_excel_btn.setFixedSize(110, 30)
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        right_buttons.addWidget(self.export_excel_btn)
        right_buttons.addStretch()

        # Üst panel bileşenlerini ekle
        upper_layout.addLayout(left_buttons)
        upper_layout.addLayout(center_layout, 1)  # 1 stretch factor
        upper_layout.addLayout(right_buttons)

        # Ana splitter
        main_splitter = QSplitter(Qt.Vertical)
        main_splitter.setHandleWidth(5)

        # Üst panel splitter'a ekle
        upper_panel.setLayout(upper_layout)
        main_splitter.addWidget(upper_panel)

        # Alt panel için splitter
        lower_splitter = QSplitter(Qt.Horizontal)

        # Sol taraf (PDF içeriği)
        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_label = QLabel('🔻 Takbis''ler''🔻')
        self.raw_text = QTextEdit()
        self.raw_text.setReadOnly(True)
        left_layout.addWidget(left_label)
        left_layout.addWidget(self.raw_text)
        left_container.setLayout(left_layout)

        # Sağ taraf (Analiz sonucu)
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_label = QLabel('🔻 Analiz Sonucu 🔻')
        self.result_text = QTextEdit()
        self.result_text.setAcceptRichText(True)

        # Font ayarları
        font = QFont()
        font.setPointSize(11)
        self.result_text.setFont(font)

        right_layout.addWidget(right_label)
        right_layout.addWidget(self.result_text)
        right_container.setLayout(right_layout)

        lower_splitter.addWidget(left_container)
        lower_splitter.addWidget(right_container)

        # Alt splitter'ı ana splitter'a ekle
        main_splitter.addWidget(lower_splitter)

        # Splitter stillerini ayarla
        main_splitter.setStyleSheet("""
            QSplitter::handle {
                background-color: #CCCCCC;
                border: 1px solid #999999;
                border-radius: 1px;
                margin: 1px;
            }
            QSplitter::handle:horizontal {
                width: 1px;
            }
            QSplitter::handle:vertical {
                height: 1px;
            }
            QSplitter::handle:hover {
                background-color: #BBBBBB;
            }
        """)

        # Splitter oranlarını ayarla
        main_splitter.setSizes([300, 500])   # Üst ve alt panel oranları
        lower_splitter.setSizes([150, 800])  # Sol ve sağ panel oranları

        # Ana layout'a splitter'ı ekle
        layout.addWidget(main_splitter)

        # Durum çubuğu
        self.statusBar().showMessage('Hazır ...')

        # Başlangıçta mevcut kayıtları göster
        self.load_database_records()
        self.raporla()

    def open_help_file(self):
        # Yardım dosyasının adı ve yolu
        help_file_name = "help.htm"
        help_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), help_file_name)

        # Dosyanın var olup olmadığını kontrol et
        if os.path.exists(help_file_path):
            try:
                # HTML dosyasını varsayılan tarayıcıda aç
                webbrowser.open(help_file_path)
            except Exception as e:
                # Hata durumunda kullanıcıyı bilgilendir
                QMessageBox.critical(self, "Hata", f"Yardım dosyası açılırken bir hata oluştu:\n{str(e)}")
        else:
            # Dosya bulunamazsa kullanıcıya uyarı ver
            QMessageBox.warning(self, "Uyarı", "Yardım dosyası bulunamadı.\nLütfen 'help.htm' dosyasının uygulama dizininde olduğundan emin olun.")

    def setup_tree_context_menu(self):
        """ QTreeWidget sağ tıklama menüsü ayarla """
        self.tree_widget.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree_widget.customContextMenuRequested.connect(self.show_tree_context_menu)

    def show_tree_context_menu(self, position):
        item = self.tree_widget.itemAt(position)
        if item is None:
            return
            
        parsel_no = item.text(0)
        try:
            if "/" in parsel_no:
                ada, parsel = parsel_no.split("/")
                baslikkk_ = f"☞ {ada} Ada {parsel} Parsel, {item.text(1)} ☜"
            else:
                baslikkk_ = f"{parsel_no} {item.text(1)}"
        except Exception as e:
            baslikkk_ = f"{parsel_no} {item.text(1)}"

        menu = QMenu(self)
        baslikkk = menu.addAction(baslikkk_)
        menu.addSeparator()
        raporla_action = menu.addAction("Raporla")
        goruntule_action = menu.addAction("Görüntüle")
        hisse_tablosu_action = menu.addAction("Hisse Tablosu Çıkart")
        menu.addSeparator()
        sil_action = menu.addAction("❌ Sil")
        menu.addSeparator()
        action = menu.exec_(self.tree_widget.viewport().mapToGlobal(position))

        if action == raporla_action:
            self.raporla_item(item)
        elif action == goruntule_action:
            self.goruntule_item(item)
        elif action == hisse_tablosu_action:
            tasinmaz_kimlik = item.text(5)  # İndeks 5'e güncellendi
            self.export_to_excel(tasinmaz_kimlik)
        elif action == sil_action:
            self.sil_item(item)

    def on_tree_double_clicked(self, item, column):
        """TreeWidget'da bir öğeye çift tıklandığında görüntüleme işlemini başlat"""
        self.goruntule_item(item)

    def raporla_item(self, item):
        """Seçilen taşınmazın detaylı raporunu oluştur"""
        tasinmaz_no = item.text(5)  # Taşınmaz numarası
    
        try:
            inceleme = TakbisInceleme()
            conn = sqlite3.connect("veritabani.db")
            cursor = conn.cursor()
        
            # Temel taşınmaz bilgilerini al                                                                             
            cursor.execute("""
                SELECT t.tasinmaz_no, tt.tapu_tarih 
                FROM tasinmaz t 
                JOIN takbis_tarih tt ON t.tasinmaz_no = tt.tasinmaz_kimlik 
                WHERE t.tasinmaz_no = ?
            """, (tasinmaz_no,))
            kayit = cursor.fetchone()
        
            if not kayit:
                QMessageBox.warning(self, "Uyarı", "Taşınmaz kaydı bulunamadı.")
                return
        
            ada, parsel, zemin_tipi, bb_detay = inceleme.get_tasinmaz_detay(cursor, tasinmaz_no)
            tarih, saat = inceleme.format_datetime(kayit[1])
        
            # Rapor başlığı ve giriş
            rapor = f"<b>{ada} ADA / {parsel} PARSEL"
            if bb_detay:
                rapor += f" - {bb_detay}"
            rapor += " TAKYİDATLARI</b><br><br>"
        
            rapor += (f"TKGM Web-Tapu portaldan elektronik ortamda {tarih} tarih ve saat {saat} itibarıyla alınan ve "
                     f"rapor ekinde yer alan Tapu Kayıt Belgesine göre <b>{ada}</b> Ada <b>{parsel}</b> Parsel "
                     f"nolu taşınmaz üzerinde aşağıda yer alan takyidat bulunmaktadır.<br><br>")
        
            # Muhdesat bilgileri
            muhdesat_kayitlar = inceleme.get_muhdesat_bilgileri(cursor, tasinmaz_no)
            if muhdesat_kayitlar:
                rapor += "<b>Muhdesat Bilgileri Hanesinde;</b><br>"
                for kayit in muhdesat_kayitlar:
                    formatted_line = inceleme.format_muhdesat_line(kayit)
                    rapor += formatted_line + "<br>"
                rapor += "<br>"
        
            # Eklenti bilgileri
            eklenti_kayitlar = inceleme.get_eklenti_bilgileri(cursor, tasinmaz_no)
            if eklenti_kayitlar:
                rapor += "<b>Eklenti Bilgileri Hanesinde;</b><br>"
                for kayit in eklenti_kayitlar:
                    formatted_line = inceleme.format_muhdesat_line(kayit)
                    rapor += formatted_line + "<br>"
                rapor += "<br>"
        
            # Şerh beyan bilgileri
            serh_beyan = inceleme.get_serh_beyan_bilgileri(cursor, tasinmaz_no)
            rapor += "<b>Taşınmaza Ait Şerh Beyan İrtifak Bilgileri hanesinde;</b><br>"
            if serh_beyan:
                for kayit in serh_beyan:
                    rapor += inceleme.format_line(kayit) + "<br>"
            else:
                rapor += "- Herhangi bir takyidat bulunmamaktadır.<br>"
            rapor += "<br>"
        
            # Teferruat bilgileri
            teferruat_kayitlar = inceleme.get_teferruat_bilgileri(cursor, tasinmaz_no)
            if teferruat_kayitlar:
                rapor += "<b>Teferruat Bilgileri hanesinde;</b><br>"
                for kayit in teferruat_kayitlar:
                    formatted_line = inceleme.format_teferruat_line(kayit)
                    rapor += formatted_line + "<br>"
                rapor += "<br>"
        
            # Mülkiyete ait şerh beyan bilgileri
            mulkiyet_kayitlar = inceleme.get_mulkiyet_serh_beyan_bilgileri(cursor, tasinmaz_no)
            rapor += "<b>Mülkiyete Ait Şerh Beyan İrtifak Bilgileri hanesinde;</b><br>"
            if mulkiyet_kayitlar:
                for kayit in mulkiyet_kayitlar:
                    formatted_line = inceleme.format_mulkiyet_line(kayit)
                    rapor += formatted_line + "<br>"
            else:
                rapor += "- Herhangi bir takyidat bulunmamaktadır.<br>"
            rapor += "<br>"
        
            # İpotek bilgileri
            ipotek = inceleme.get_ipotek_bilgileri(cursor, tasinmaz_no)
            rapor += "<b>Mülkiyete Ait Rehin Bilgileri Hanesinde;</b><br>"
            if ipotek:
                for kayit in ipotek:
                    rapor += inceleme.format_ipotek_line(kayit) + "<br>"
            else:
                rapor += "- Herhangi bir takyidat bulunmamaktadır.<br>"
        
            self.result_text.setHtml(rapor)
        
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Rapor oluşturma hatası: {str(e)}")
            logging.error(f"Rapor oluşturma hatası: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()

    def goruntule_item(self, item):
        """Seçilen taşınmazın özet bilgilerini göster"""
        tasinmaz_no = item.text(5)

        try:
            conn = sqlite3.connect("veritabani.db")
            cursor = conn.cursor()

            # Taşınmaz temel bilgileri ve ek bilgiler
            cursor.execute("""
                SELECT t.il_ilce, t.mahalle, t.ada_parsel, t.zemintipi, 
                       t.ana_tasinmaz_nitelik, t.bb_nitelik, t.blok_kat_girisi_bbno,
                       t.cilt_sayfa_no, t.at_yuzolcum, t.arsa_pay_payda,
                       tt.tapu_tarih
                FROM tasinmaz t
                LEFT JOIN takbis_tarih tt ON t.tasinmaz_no = tt.tasinmaz_kimlik
                WHERE t.tasinmaz_no = ?
            """, (tasinmaz_no,))
            tasinmaz = cursor.fetchone()

            if not tasinmaz:
                QMessageBox.warning(self, "Uyarı", "Taşınmaz bilgisi bulunamadı.")
                return

            # Malik/Hissedar sayısını al
            cursor.execute("""
                SELECT COUNT(*) 
                FROM tapu_verileri 
                WHERE tasinmaz_kimlik = ? 
                AND takyidat_baslik = 'MÜLKİYET BİLGİLERİ'
                AND baslikontrol = 'HAYIR'
            """, (tasinmaz_no,))
            malik_sayisi = cursor.fetchone()[0]

            # Başlıkları ve kayıt sayılarını al
            cursor.execute("""
                SELECT DISTINCT b.baslik, COUNT(tv.rowid) as kayit_sayisi,
                       CASE WHEN b.baslik = 'MÜLKİYETE AİT REHİN BİLGİLERİ' THEN
                           (SELECT COUNT(DISTINCT id) FROM ipotek_verileri WHERE tasinmaz_kimlik = b.tasinmaz_kimlik)
                       ELSE COUNT(tv.rowid) END as gercek_sayi
                FROM baslik_bilgileri b
                LEFT JOIN tapu_verileri tv ON tv.tasinmaz_kimlik = b.tasinmaz_kimlik 
                    AND tv.takyidat_baslik = b.baslik
                    AND tv.baslikontrol = 'HAYIR'
                WHERE b.tasinmaz_kimlik = ?
                    AND b.baslik != 'TAPU KAYIT BİLGİSİ'
                GROUP BY b.baslik
                ORDER BY b.baslik
            """, (tasinmaz_no,))
            baslik_sayilari = cursor.fetchall()

            # Tarih ve saat formatla
            tarih_saat = ""
            if tasinmaz[10]:  # tapu_tarih
                tarih_parts = tasinmaz[10].split('-')
                if len(tarih_parts) >= 4:
                    tarih = f"{tarih_parts[0]}-{tarih_parts[1]}-{tarih_parts[2]}"
                    saat = tarih_parts[3]
                    tarih_saat = f"{tarih} / {saat}"

            # Özet rapor oluştur
            ozet = "<b>TAŞINMAZ ÖZET BİLGİLERİ</b><br><br>"
            if tarih_saat:
                ozet += f"<b>Takbis Belgesi Alındığı Tarih/Saat:</b> {tarih_saat}<br>"
            ozet += f"<b>İl/İlçe:</b> {tasinmaz[0]}<br>"
            ozet += f"<b>Mahalle:</b> {tasinmaz[1]}<br>"
            ozet += f"<b>Ada/Parsel:</b> {tasinmaz[2]}<br>"
            ozet += f"<b>Cilt/Sayfa No:</b> {tasinmaz[7]}<br>"

            # Yüzölçüm bilgisi
            if tasinmaz[8]:  # at_yuzolcum
                yuzolcum = str(tasinmaz[8]).replace('.', ',')
                ozet += f"<b>Yüzölçümü:</b> {yuzolcum} m²<br>"

            ozet += f"<b>Zemin Tipi:</b> {tasinmaz[3]}<br>"

            if tasinmaz[3] in ["KatIrtifaki", "KatMulkiyeti"]:
                ozet += f"<b>Bağımsız Bölüm Niteliği:</b> {tasinmaz[5]}<br>"
                if tasinmaz[9]:  # arsa_pay_payda
                    ozet += f"<b>Arsa Pay/Payda:</b> {tasinmaz[9]}<br>"
                if tasinmaz[6]:  # BB bilgisi
                    bb_parts = tasinmaz[6].split('/')
                    bb_info = []
                    if bb_parts[0] and bb_parts[0] != "-":  # Blok
                        bb_info.append(f"Blok: {bb_parts[0]}")
                    if len(bb_parts) > 1:  # Kat
                        bb_info.append(f"Kat: {bb_parts[1]}")
                    if len(bb_parts) > 2 and bb_parts[2] != "-":  # Giriş
                        bb_info.append(f"Giriş: {bb_parts[2]}")
                    if len(bb_parts) > 3:  # BB No
                        bb_info.append(f"BB No: {bb_parts[3]}")
                    ozet += f"<b>Bağımsız Bölüm Bilgisi:</b> {' | '.join(bb_info)}<br>"
            else:
                ozet += f"<b>Ana Taşınmaz Niteliği:</b> {tasinmaz[4]}<br>"

            ozet += "<br><b>MALİK BİLGİLERİ</b><br>"
            ozet += f"<b>Malik/Hissedar Sayısı:</b> {malik_sayisi} adet<br>"

            ozet += "<br><b>TAKYİDAT ÖZETİ</b><br>"
        
            has_takyidat = False
            for baslik, _, gercek_sayi in baslik_sayilari:
                if gercek_sayi > 0:
                    has_takyidat = True
                    # Başlık ismini daha okunabilir hale getir
                    if baslik == "MÜLKİYETE AİT REHİN BİLGİLERİ":
                        ozet += f"<b>Rehin (İpotek) Bilgileri:</b> {gercek_sayi} adet<br>"
                    else:
                        baslik_gosterim = baslik.title().replace("Bilgileri", "")
                        ozet += f"<b>{baslik_gosterim}:</b> {gercek_sayi} adet<br>"

            if not has_takyidat:
                ozet += "<br><i>Bu taşınmaz üzerinde herhangi bir takyidat bulunmamaktadır.</i>"
        
            self.result_text.setHtml(ozet)

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Görüntüleme hatası: {str(e)}")
            logging.error(f"Görüntüleme hatası: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()

    def sil_item(self, item):
        """ Veritabanından taşınmazı sil """
        tasinmaz_no = item.text(5)
        if not tasinmaz_no:
            QMessageBox.warning(self, "Hata", "Silinecek taşınmaz numarası bulunamadı.")
            return

        # Silme işlemi için SQL sorgusu
        conn = sqlite3.connect("veritabani.db")
        cursor = conn.cursor()
        tablolar = ['ipotek_verileri', 'takbis_tarih', 'tapu_verileri', 'koordinat_bilgileri_ext', 'baslik_bilgileri']

        try:
            for tablo in tablolar:
                cursor.execute(f"DELETE FROM  {tablo} WHERE tasinmaz_kimlik = ?", (tasinmaz_no,))
                
            cursor.execute(f"DELETE FROM  tasinmaz WHERE tasinmaz_no = ?", (tasinmaz_no,))                    
            conn.commit()
            QMessageBox.information(self, "Başarılı", f"Taşınmaz {tasinmaz_no} başarıyla silindi.")

            # QTreeWidget güncelle
            self.load_database_records()
            self.raporla()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Silme işlemi sırasında hata oluştu: {str(e)}")
        finally:
            conn.close()

    def veritabani_olustur(self):
        """Veritabanı ve gerekli tabloları oluşturur"""
        try:
            conn = sqlite3.connect("veritabani.db")
            cursor = conn.cursor()

            # Tabloları oluştur
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS baslik_bilgileri (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tasinmaz_kimlik TEXT,
                    baslik TEXT,
                    sayfa_no INTEGER,
                    y_koordinat REAL,
                    auto_detected BOOLEAN DEFAULT FALSE,
                    tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    baslik_deger INTEGER
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS ipotek_verileri (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tapu_tarih TEXT,
                    tasinmaz_kimlik TEXT,
                    sayfa_no INTEGER,
                    alacakli TEXT,
                    musterek_mi TEXT,
                    borc TEXT,
                    faiz TEXT,
                    derece_sira TEXT,
                    sure TEXT,
                    tesis_tarih TEXT,
                    tasinmaz TEXT,
                    hisse_pay_payda TEXT,
                    borclu_malik TEXT,
                    sn_bilgisi TEXT,
                    malik_borc TEXT,
                    tescil_tarih TEXT,
                    terkin TEXT
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS koordinat_bilgileri_ext (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tapu_rowid INTEGER,
                    hucreno_1_deger TEXT,
                    y_koordinat REAL,
                    tasinmaz_kimlik TEXT,
                    sayfano INTEGER,
                    satirno INTEGER,
                    baslik_deger INTEGER
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS takbis_tarih (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tapu_tarih TEXT,
                    tasinmaz_kimlik TEXT,
                    kayit_tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS tapu_verileri (
                    sayfano INTEGER NOT NULL,
                    satirno INTEGER NOT NULL,
                    hucreno_1 TEXT,
                    hucreno_2 TEXT,
                    hucreno_3 TEXT,
                    hucreno_4 TEXT,
                    hucreno_5 TEXT,
                    hucreno_6 TEXT,
                    hucreno_7 TEXT,
                    hucreno_8 TEXT,
                    hucreno_9 TEXT,
                    hucreno_10 TEXT,
                    hucreno_11 TEXT,
                    baslikontrol TEXT DEFAULT 'HAYIR',
                    tasinmaz_kimlik TEXT,
                    takyidat_baslik TEXT,
                    takyidat_1 TEXT,
                    takyidat_2 TEXT,
                    takyidat_3 TEXT
                )
            """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS tasinmaz (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tasinmaz_no TEXT UNIQUE,
                    zemintipi TEXT,
                    il_ilce TEXT,
                    kurum_adi TEXT,
                    mahalle TEXT,
                    mevki TEXT,
                    cilt_sayfa_no TEXT,
                    kayitdurumu TEXT,
                    ada_parsel TEXT,
                    at_yuzolcum TEXT,
                    bb_nitelik TEXT,
                    bb_brüt_yuzolcum TEXT,
                    bb_net_yuzolcum TEXT,
                    blok_kat_girisi_bbno TEXT,
                    arsa_pay_payda TEXT,
                    ana_tasinmaz_nitelik TEXT,
                    tapu_tarih TEXT,
                    baslik_islem_yapildi BOOLEAN DEFAULT FALSE
                )
            """)

            conn.commit()
            return True

        except Exception as e:
            print(f"Veritabanı oluşturma hatası: {str(e)}")
            return False

        finally:
            if conn:
                conn.close()

    def load_database_records(self):
        """Veritabanından kayıtları yükle ve TreeWidget'a ekle"""
        # Önce mevcut içeriği temizle
        self.tree_widget.clear()
    
        # TreeWidget başlıklarını güncelle
        self.tree_widget.setHeaderLabels(['Ada/Parsel', 'Nitelik', 'BB/Blok Bilgisi', 'İl/İlçe', 'Mahalle', 'Taşınmaz No'])

        # Veritabanı dosyasının varlığını kontrol et
        if not os.path.exists("veritabani.db"):
            self.veritabani_olustur()            
            return
    
        try:
            conn = sqlite3.connect("veritabani.db")
            cursor = conn.cursor()
    
            # Tasinmaz tablosunun varlığını kontrol et
            cursor.execute("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name='tasinmaz'
            """)
    
            if not cursor.fetchone():
                self.veritabani_olustur()   
                return
        
            # Var olan BB'li kayıt sayısını kontrol et
            cursor.execute("""
                SELECT COUNT(*) FROM tasinmaz 
                WHERE blok_kat_girisi_bbno IS NOT NULL 
                AND blok_kat_girisi_bbno != ''
            """)
            bb_count = cursor.fetchone()[0]
    
            # Var olan BB'siz kayıt sayısını kontrol et
            cursor.execute("""
                SELECT COUNT(*) FROM tasinmaz 
                WHERE blok_kat_girisi_bbno IS NULL 
                OR blok_kat_girisi_bbno = ''
            """)
            non_bb_count = cursor.fetchone()[0]
       
            if bb_count > 0:
                bb_parent = QTreeWidgetItem(self.tree_widget)
                bb_parent.setText(0, f"Bağımsız Bölümlü Taşınmazlar ({bb_count} adet)")
                bb_parent.setExpanded(True)
        
                cursor.execute("""
                    SELECT ada_parsel, bb_nitelik, blok_kat_girisi_bbno, 
                           il_ilce, mahalle, tasinmaz_no 
                    FROM tasinmaz 
                    WHERE blok_kat_girisi_bbno IS NOT NULL 
                    AND blok_kat_girisi_bbno != ''
                    ORDER BY ada_parsel
                """)
    
                for row in cursor.fetchall():
                    item = QTreeWidgetItem(bb_parent)
                    item.setText(0, str(row[0]))  # Ada/Parsel
                    item.setText(1, str(row[1]))  # BB Nitelik
                    item.setText(2, str(row[2]))  # BB/Blok Bilgisi
                    item.setText(3, str(row[3]))  # İl/İlçe
                    item.setText(4, str(row[4]))  # Mahalle
                    item.setText(5, str(row[5]))  # Taşınmaz No
            
            if non_bb_count > 0:
                non_bb_parent = QTreeWidgetItem(self.tree_widget)
                non_bb_parent.setText(0, f"Ana Taşınmazlar ({non_bb_count} adet)")
                non_bb_parent.setExpanded(True)
        
                cursor.execute("""
                    SELECT ada_parsel, ana_tasinmaz_nitelik, 
                           il_ilce, mahalle, tasinmaz_no 
                    FROM tasinmaz 
                    WHERE blok_kat_girisi_bbno IS NULL 
                    OR blok_kat_girisi_bbno = ''
                    ORDER BY ada_parsel
                """)
    
                for row in cursor.fetchall():
                    item = QTreeWidgetItem(non_bb_parent)
                    item.setText(0, str(row[0]))  # Ada/Parsel
                    item.setText(1, str(row[1]))  # Ana Taşınmaz Nitelik
                    item.setText(3, str(row[2]))  # İl/İlçe
                    item.setText(4, str(row[3]))  # Mahalle
                    item.setText(5, str(row[4]))  # Taşınmaz No
    
            # Sütun genişliklerini ayarla
            for i in range(6):  # Kolon sayısını 6'ya çıkardık
                self.tree_widget.resizeColumnToContents(i)
        
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Veritabanı yükleme hatası: {str(e)}")
        finally:
            if 'conn' in locals():
                conn.close()

    def select_multiple_files(self):
        """Birden fazla PDF dosyası seçme"""
        try:
            # Masaüstü yolunu al
            desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
            if not os.path.exists(desktop_path):
                # Türkçe Windows sistemleri için alternatif yol
                desktop_path = os.path.join(os.path.expanduser('~'), 'Masaüstü')
        
            # Masaüstü yolu varsa ve erişilebilirse
            if os.path.exists(desktop_path) and os.access(desktop_path, os.R_OK):
                files, _ = QFileDialog.getOpenFileNames(
                    self,
                    "PDF Dosyalarını Seç",
                    desktop_path,  # Masaüstü yolunu kullan
                    "PDF Dosyaları (*.pdf)"
                )
            else:
                # Masaüstüne erişilemezse varsayılan yolu kullan
                files, _ = QFileDialog.getOpenFileNames(
                    self,
                    "PDF Dosyalarını Seç",
                    "",
                    "PDF Dosyaları (*.pdf)"
                )

        except Exception as e:
            # Herhangi bir hata durumunda varsayılan yolu kullan
            logging.error(f"Masaüstü yolu açılırken hata: {str(e)}")
            files, _ = QFileDialog.getOpenFileNames(
                self,
                "PDF Dosyalarını Seç",
                "",
                "PDF Dosyaları (*.pdf)"
            )

        if files:
            self.file_list = files
            # Seçilen dosya sayısını göster
            self.statusBar().showMessage(f'{len(files)} adet PDF dosyası seçildi')
            # Seçilen dosyaları raw_text'e listele
            file_list_text = "Seçilen Dosyalar:<br>"
            for i, file in enumerate(files, 1):
                file_list_text += f"{i}. {os.path.basename(file)}<br>"
            self.raw_text.setHtml(file_list_text)

    def start_batch_processing(self):
        """Toplu işleme başlat"""
        if not hasattr(self, 'file_list') or not self.file_list:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce PDF dosyalarını seçin.")
            return
        
        # İşlem başlamadan önce onay al
        reply = QMessageBox.question(
            self,
            'Toplu İşlem',
            f'{len(self.file_list)} adet dosya işlenecek. Devam etmek istiyor musunuz?',
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
    
        if reply == QMessageBox.Yes:
            self.process_files()

    def save_report(self):
        """Raporu kaydet"""
        if not self.result_text.toPlainText():
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek rapor bulunamadı.")
            return

        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Raporu Kaydet",
            "",
            "Text Dosyaları (*.txt);;Tüm Dosyalar (*)"
        )
        
        if file_name:
            try:
                with open(file_name, 'w', encoding='utf-8') as file:
                    file.write(self.result_text.toPlainText())
                self.statusBar().showMessage(f'Rapor kaydedildi: {file_name}')
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Kaydetme hatası: {str(e)}")

    def clear_text(self):
        """Metin alanlarını ve veritabanını temizle"""
        # Metin alanlarını temizle
        self.result_text.clear()
        try:
            # Kullanıcıdan onay al
            cevap = QMessageBox.question(
                self,
                "Veritabanı Temizleme",
                "Tüm veritabanı kayıtları silinecek. Emin misiniz?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if cevap == QMessageBox.Yes:
                # Log dosyalarını temizle
                try:
                    log_dir = "logs"
                    if os.path.exists(log_dir):
                        # Log dosyalarını al ve tarihe göre sırala
                        log_files = []
                        for file in os.listdir(log_dir):
                            if file.startswith('tapu_analiz_') and file.endswith('.log'):
                                file_path = os.path.join(log_dir, file)
                                log_files.append((file_path, os.path.getmtime(file_path)))
                    
                        # Tarihe göre sırala (en yeni en üstte)
                        log_files.sort(key=lambda x: x[1], reverse=True)
                    
                        # En yeni dosyayı hariç tut, diğerlerini sil
                        if len(log_files) > 1:
                            for file_path, _ in log_files[1:]:
                                try:
                                    os.remove(file_path)
                                except:
                                    pass
                except:
                    pass  # Log temizleme hatalarını sessizce geç

                # Veritabanı bağlantısı
                conn = sqlite3.connect("veritabani.db")
                cursor = conn.cursor()

                # Tablo listesi
                tablolar = [
                    'ipotek_verileri',
                    'sqlite_sequence',
                    'takbis_tarih',
                    'tasinmaz',
                    'tapu_verileri',
                    'koordinat_bilgileri_ext',
                    'baslik_bilgileri'
                ]

                # Her tabloyu temizle
                for tablo in tablolar:
                    try:
                        cursor.execute(f"DELETE FROM {tablo}")
                        logging.info(f"{tablo} tablosu temizlendi")
                    except sqlite3.OperationalError as e:
                        logging.warning(f"{tablo} tablosu temizlenirken hata: {str(e)}")
                        continue

                conn.commit()
                conn.close()

                # TreeWidget'ı temizle ve güncelle
                self.tree_widget.clear()  # Önce mevcut içeriği temizle
                self.load_database_records()  # Sonra yeniden yükle

                # Raporu güncelle
                self.raporla()

                self.statusBar().showMessage('Tüm veriler temizlendi')
                QMessageBox.information(self, "Başarılı", "Tüm veritabanı kayıtları silindi.")

            else:
                self.statusBar().showMessage('Temizleme iptal edildi')

        except Exception as e:
            logging.error(f"Veritabanı temizleme hatası: {str(e)}")
            QMessageBox.critical(self, "Hata", f"Temizleme sırasında hata oluştu: {str(e)}")

    def raporla(self):        
        """Veritabanındaki taşınmaz kayıtlarını raporlar"""
        try:
            # Veritabanı bağlantısı
            conn = sqlite3.connect("veritabani.db")
            cursor = conn.cursor()
        
            # Önce tablo var mı kontrol et
            cursor.execute("""
                SELECT name FROM sqlite_master 
                WHERE type='table' AND name='tapu_verileri'
            """)
        
            table_exists = cursor.fetchone() is not None
        
            if table_exists:
                # Tablo varsa kayıtları say
                cursor.execute("SELECT COUNT(DISTINCT tasinmaz_kimlik) FROM tapu_verileri")
                kayit_sayisi = cursor.fetchone()[0]
                if kayit_sayisi >0:
                    rapor_metni = f"{kayit_sayisi} Adet Takbis Kaydı Mevcut\n"
                else:
                    rapor_metni = f"Henüz Takbis Kaydı Yok\n"
                    
            else:
                # Tablo yoksa bilgilendirme mesajı
                rapor_metni = "Henüz Takbis Kaydı Yok\n"
        
            # Metni göster            
            self.raw_text.setHtml(rapor_metni)    
        except Exception as e:
            logging.debug(f"Raporlama bilgi mesajı: {str(e)}")
            self.result_text.setHtml("Henüz Takbis Kaydı Yok\n")
        finally:
            if 'conn' in locals() and conn:
                conn.close()

    def analyze_takbis(self):
        """Takbisleri incele ve sonucu göster"""
        try:
            inceleme = TakbisInceleme()
            rapor = inceleme.incele()
            # result_text'i temizle
            self.result_text.clear()
            # Yeni raporu göster
            self.result_text.setHtml(rapor)
            self.statusBar().showMessage('Takbis inceleme tamamlandı')
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"İnceleme sırasında hata oluştu: {str(e)}")
                
    def show_multi_format(self):        
        """Raporu göster"""
        try:
            inceleme = CokluInceleme()
            report = inceleme.incele()
        
            # result_text'i temizle
            self.result_text.clear()
            # Raporu göster
            self.result_text.setHtml(report)
            self.statusBar().showMessage('Rapor oluşturuldu')

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Rapor oluşturulurken hata: {str(e)}")

    def process_files(self):
        """Seçilen dosyaları toplu işle (10'arlı gruplar halinde, veritabanı temizliğiyle)"""
        if not hasattr(self, 'file_list') or not self.file_list:
            QMessageBox.warning(self, "Uyarı", "Lütfen önce PDF dosyalarını seçin.")
            return

        total_files = len(self.file_list)
        db_path = "veritabani.db"

        # Ext koordinat tablosunu oluştur
        self.table_analyzer.create_ext_koordinat_table(db_path)

        # İlerleme dialog'unu oluştur
        self.progress_dialog = ProcessingDialog(total_files, self)
        self.progress_dialog.show()
        self.progress_dialog.start_processing()

        successful_files = []
        failed_files = []

        try:
            # Dosyaları 10'arlı gruplar halinde işle
            for i in range(0, total_files, 1):
                batch_files = self.file_list[i:i + 1]  # 10 dosyadan oluşan bir grup oluştur
                for index, file_path in enumerate(batch_files, i + 1):  # index'i doğru hesapla
                    try:
                        # İlerlemeyi güncelle
                        self.progress_dialog.update_progress(file_path, index)
                        QApplication.processEvents()

                        self.current_file = file_path

                        # PDF'i aç ve analiz et
                        with pdfplumber.open(file_path) as pdf:
                            text = pdf.pages[0].extract_text()

                            # Doğrulama işlemleri
                            if not self.table_analyzer.tapu_belgesi_kontrol(text):
                                raise ValueError("Geçerli bir tapu belgesi değil")

                            tasinmaz_no = self.table_analyzer.tasinmaz_kimlik_no_al(text)
                            if not tasinmaz_no:
                                raise ValueError("Taşınmaz Kimlik No bulunamadı")

                            if self.table_analyzer.tasinmaz_kayit_kontrol(tasinmaz_no, db_path):
                                self.progress_dialog.add_detail(
                                    f"ATLANDI - {os.path.basename(file_path)}: "
                                    f"Bu taşınmaz (Kimlik No: {tasinmaz_no}) zaten kayıtlı"
                                )
                                continue

                            # Fitz ile koordinatları çıkar ve kaydet
                            self.table_analyzer.extract_fitz_coordinates(file_path, tasinmaz_no)

                            # İpotek verilerini çıkar
                            ipotek_extractor = IpotekKoordinatExtractor(db_path)
                            ipotek_sonuc = ipotek_extractor.extract_from_pdf(file_path)

                            # Enhanced analizi kullan
                            analiz_basarili = enhanced_analyze_pdf(file_path, db_path, tasinmaz_no)

                            if not analiz_basarili:
                                raise ValueError("PDF analizi başarısız oldu")

                            # Tapu düzenleme işlemleri
                            processor = TapuProcessor(db_path)
                            if processor.process_all():
                                # Tesis tarih işlemlerini yap
                                tesis_processor = TesisProcessor(db_path)
                                tesis_sonuc = tesis_processor.process_all()

                                # FitzTapu analizini yap
                                try:
                                    fitz_analyzer = FitzTapuAnalyzer(db_path, tasinmaz_no)
                                    analysis_result = fitz_analyzer.analyze_pdf(file_path)

                                    if analysis_result['status'] == 'success':
                                        self.progress_dialog.add_detail(
                                            f"BAŞLIK ANALİZİ - {os.path.basename(file_path)}: "
                                            f"{len(analysis_result['headers'])} başlık tespit edildi"
                                        )
                                    else:
                                        self.progress_dialog.add_detail(
                                            f"BAŞLIK ANALİZ UYARISI - {os.path.basename(file_path)}: "
                                            f"{analysis_result['message']}"
                                        )
                                except Exception as e:
                                    self.progress_dialog.add_detail(
                                        f"BAŞLIK ANALİZ HATASI - {os.path.basename(file_path)}: "
                                        f"{str(e)}"
                                    )

                                if tesis_sonuc:
                                    if ipotek_sonuc:
                                        success_message = f"BAŞARILI - {os.path.basename(file_path)}: Veriler ve ipotek bilgileri kaydedildi. (Taşınmaz No: {tasinmaz_no})"
                                    else:
                                        success_message = f"BAŞARILI - {os.path.basename(file_path)}: Veriler ve tesis bilgileri kaydedildi, ipotek bilgisi bulunamadı. (Taşınmaz No: {tasinmaz_no})"
                                else:
                                    success_message = f"KISMEN BAŞARILI - {os.path.basename(file_path)}: Veriler kaydedildi fakat tesis işlemleri başarısız. (Taşınmaz No: {tasinmaz_no})"

                                self.progress_dialog.add_detail(success_message)
                                successful_files.append(file_path)
                            else:
                                raise ValueError("Veri işleme hatası")

                    except Exception as e:
                        failed_files.append((file_path, str(e)))
                        self.progress_dialog.add_detail(
                            f"HATA - {os.path.basename(file_path)}: {str(e)}"
                        )

                # Her 1 dosyadan sonra veritabanı temizliği ve arayüz güncellemeleri
                try:
                    #Veritabanı gereksiz kayıt temizleme
                    processor = TapuProcessor(db_path)
                    baslik_sonuc = processor.add_baslik_deger_columns()
                    processor.update_missing_headers()

                    # Yeni koordinat düzeltme işlemini ekle
                    koordinat_sonuc = processor.update_coordinate_assignments()                
                    if koordinat_sonuc:
                        self.progress_dialog.add_detail(
                            "\n✓"
                        )
                    else:
                        self.progress_dialog.add_detail(
                            "\nUYARI - Koordinat düzeltme işleminde sorun oluştu"
                        )                    

                except Exception as e:
                    self.progress_dialog.add_detail(f"\nHATA - Veritabanı temizliği sırasında: {str(e)}")

                #Tree widgeti güncelle
                self.load_database_records()
                #Arayüzü de raporla metodu ile güncelle
                self.raporla()

            # İşlem sonuç özeti
            total_processed = len(successful_files)
            total_failed = len(failed_files)

            summary = f"""
            İşlem Tamamlandı
            ----------------
            Toplam Dosya: {total_files}
            Başarılı: {total_processed}
            Başarısız: {total_failed}
            """

            self.progress_dialog.add_detail("\n" + summary)

            # TreeWidget'ı güncelle
            self.load_database_records()

            self.raporla()

            self.progress_dialog.finish_processing()

            QMessageBox.information(
                self,
                "İşlem Tamamlandı",
                f"Toplam {total_files} dosyadan:\n"
                f"Başarılı :{total_processed}\n"
                f"Başarısız:{total_failed}"
            )
            
            tapu_processor = TapuProcessor("veritabani.db")
            success, message = tapu_processor.delete_empty_cells_with_yevmiye()
            if success:
                print(message)
            else:
                print("Hata:", message)

        except Exception as e:
            self.progress_dialog.finish_processing()
            QMessageBox.critical(self, "Hata", f"İşlem sırasında beklenmeyen hata: {str(e)}")

    def export_to_excel(self, tasinmaz_kimlik=None):     
        try:
            # Veritabanı bağlantısı
            conn = sqlite3.connect("veritabani.db")
            cursor = conn.cursor()

            # Helper fonksiyonlar
            def clean_name(adi_soyadi):
                if not adi_soyadi:
                    return ""
            
                # Filigran harflerini temizle
                filigran_harfler = ["B", "İ", "L", "G", "İ", "A", "M", "A", "Ç", "L", "I", "D", "R"]
            
                # SN numarasını temizle
                adi_soyadi = re.sub(r'\(SN:\d+L?\d*\)', '', adi_soyadi)
            
                # İki nokta formatını düzelt
                parts = adi_soyadi.split(':')
                if len(parts) == 2:
                    isim = parts[0].strip()
                    baba = parts[1].strip()
                
                    # İsim kısmından filigran harflerini temizle
                    isim_words = [word for word in isim.split() if word not in filigran_harfler]
                    isim = ' '.join(isim_words)
                
                    # Baba adı kısmından filigran harflerini temizle
                    baba_words = [word for word in baba.split() if word not in filigran_harfler]
                    baba = ' '.join(baba_words)
                
                    # Temizlenmiş formatı birleştir
                    adi_soyadi = f"{isim} {baba}"
            
                # Fazla boşlukları temizle
                adi_soyadi = ' '.join(adi_soyadi.split())
            
                return adi_soyadi.strip()

            def clean_pay_payda(pay_payda_str):
                if not pay_payda_str:
                    return '0/1'
                
                # Filigran harflerini temizle    
                text = re.sub(r'^[BİLGİAMAÇLIDR]\s*', '', str(pay_payda_str))
            
                # Tüm boşlukları kaldır
                text = ''.join(text.split())
            
                # Pay/payda formatını düzelt
                if '/' in text:
                    parts = text.split('/')
                    if len(parts) == 2:
                        # İlk kısım pay
                        pay = re.sub(r'[^\d]', '', parts[0])
                        # İkinci kısım payda + ek sayılar
                        payda = re.sub(r'[^\d]', '', parts[1])
                    
                        if pay and payda:
                            return f"{pay}/{payda}"
            
                return '0/1'

            # Ana sorguyu oluştur
            base_query = """
                SELECT DISTINCT 
                    tv.tasinmaz_kimlik,
                    tv.hucreno_2,
                    tv.hucreno_4,
                    tv.hucreno_5,
                    tv.hucreno_6,
                    t.il_ilce,
                    t.mahalle,
                    t.ada_parsel,
                    t.zemintipi,
                    t.ana_tasinmaz_nitelik,
                    t.bb_nitelik,
                    t.blok_kat_girisi_bbno
                FROM tapu_verileri tv
                LEFT JOIN tasinmaz t ON tv.tasinmaz_kimlik = t.tasinmaz_no
                WHERE tv.takyidat_baslik = 'MÜLKİYET BİLGİLERİ'
            """

            if tasinmaz_kimlik:
                base_query += " AND tv.tasinmaz_kimlik = ?"
                cursor.execute(base_query, (tasinmaz_kimlik,))
            else:
                base_query += " ORDER BY tv.tasinmaz_kimlik"
                cursor.execute(base_query)

            data = cursor.fetchall()

            if not data:
                QMessageBox.warning(self, "Uyarı", "İşlenecek kayıt bulunamadı.")
                return

            # Excel dosya adını belirle
            if tasinmaz_kimlik:
                default_filename = f"Hisse_Tablosu_{tasinmaz_kimlik}.xlsx"
            else:
                default_filename = "Hisse_Tablosu.xlsx"

            file_path, _ = QFileDialog.getSaveFileName(
                self, "Excel Dosyasını Kaydet", default_filename, 
                "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)"
            )

            if not file_path:
                return

            # Excel oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hisse Detay"

            # Stil tanımlamaları
            header_font = Font(name='Calibri', bold=True, size=11)
            cell_font = Font(name='Calibri', size=10)
            header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            alt_row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Başlık stili
            title_font = Font(name='Calibri', bold=True, size=14)
            title_alignment = Alignment(horizontal='center', vertical='center')

            # Sütun genişlikleri
            column_widths = {
                'A': 8,   # Sıra No
                'B': 40,  # Adı Soyadı
                'C': 15,  # Pay/Payda
                'D': 12,  # Pay
                'E': 12,  # Payda
                'F': 12,  # Oran
                'G': 20,  # Hisseye Düşen
                'H': 24,  # Toplam Yüzölçüm
                'I': 24,  # Yasal Değer
                'J': 24   # Mevcut Değer
            }

            # Sütun hizalamaları
            alignments = {
                'A': Alignment(horizontal='center', vertical='center'),  # Sıra No
                'B': Alignment(horizontal='left', vertical='center'),    # Adı Soyadı
                'C': Alignment(horizontal='center', vertical='center'),  # Pay/Payda
                'D': Alignment(horizontal='center', vertical='center'),  # Pay
                'E': Alignment(horizontal='center', vertical='center'),  # Payda
                'F': Alignment(horizontal='center', vertical='center'),  # Oran
                'G': Alignment(horizontal='right', vertical='center'),   # Hisseye Düşen
                'H': Alignment(horizontal='right', vertical='center'),   # Toplam Yüzölçüm
                'I': Alignment(horizontal='right', vertical='center'),   # Yasal Değer
                'J': Alignment(horizontal='right', vertical='center')    # Mevcut Değer
            }

            # Sütun genişliklerini ayarla
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            current_row = 1
            current_tasinmaz = None
            hisse_count = 0

            # Her kayıt için
            for row in data:
                tasinmaz_kimlik = row[0]
            
                # Yeni taşınmaz başlangıcı
                if current_tasinmaz != tasinmaz_kimlik:
                    # İlk kayıt değilse 2 boş satır ekle
                    if current_tasinmaz is not None:
                        current_row += 2

                    current_tasinmaz = tasinmaz_kimlik
                    hisse_count = 0

                    # Başlık oluştur
                    il_ilce = (row[5] or '').split('/')
                    if len(il_ilce) != 2:
                        il_ilce = ['', '']
            
                    mahalle = row[6] or ''
                    ada_parsel = (row[7] or '').split('/')
                    if len(ada_parsel) != 2:
                        ada_parsel = ['', '']
            
                    zemin_tipi = row[8] or ''
                    ana_nitelik = row[9] or ''
                    bb_nitelik = row[10] or ''

                    # Başlık oluştur
                    if not zemin_tipi or zemin_tipi not in ["KatIrtifaki", "KatMulkiyeti", "KatIr", "KatM"]:
                        baslik = (f"{il_ilce[0].strip().title()} İli {il_ilce[1].strip().title()} İlçesi "
                                  f"{mahalle.strip()} {ada_parsel[0].strip()} Ada {ada_parsel[1].strip()} "
                                  f"Parsel {ana_nitelik.strip()} Nitelikli Gayrimenkul Hisse Detay Tablosu")
                    else:
                        bb_bilgisi = row[11] if row[11] else ''
                        formatted_bb = ""

                        if bb_bilgisi:
                            parts = bb_bilgisi.split('/')
    
                            blok_kismi = ""
                            if parts[0].strip() and parts[0].strip() != "-":
                                blok_kismi = f"{parts[0].strip()} Blok "
    
                            kat = parts[1].strip() if len(parts) > 1 else ""
                            if kat:
                                if '+' in kat or 'BODRUM' in kat.upper():
                                    kat_text = f"{kat} Kat"
                                else:
                                    kat_text = f"{kat}. Kat"
    
                            giris_kismi = ""
                            if len(parts) > 2 and parts[2].strip() and parts[2].strip() != "-":
                                giris_kismi = f" {parts[2].strip()}. Giriş"
    
                            bb_no = parts[3].strip() if len(parts) > 3 else ""
    
                            formatted_bb = f"{blok_kismi}{kat_text}{giris_kismi} {bb_no} Nolu Bağımsız Bölüm"

                        baslik = (f"{il_ilce[0].strip().title()} İli {il_ilce[1].strip().title()} İlçesi "
                                  f"{mahalle.strip()} {ada_parsel[0].strip()} Ada {ada_parsel[1].strip()} "
                                  f"Parsel {bb_nitelik.strip()} Nitelikli {formatted_bb} Hisse Detay Tablosu")

                    # Başlık formatı
                    ws.merge_cells(f'A{current_row}:J{current_row}')
                    ws[f'A{current_row}'] = baslik
                    ws[f'A{current_row}'].font = title_font
                    ws[f'A{current_row}'].alignment = title_alignment
                    ws.row_dimensions[current_row].height = 35

                    current_row += 1

                    # Tablo başlıkları
                    headers = [
                        "Sıra No", "Adı Soyadı", "Pay/Payda", "Pay", "Payda", "Oran",
                        "Hissesine Düşen (m²)", "Toplam Yüzölçüm (m²)", 
                        "Yasal Değer (TL)", "Mevcut Değer (TL)"
                    ]
            
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.value = header
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                        cell.alignment = alignments[chr(64 + col)]

                    current_row += 1

                # Hisse verilerini işle
                try:
                    hisse_count += 1

                    # Adı soyadı temizleme
                    adi_soyadi = clean_name(row[1])

                    # Pay/payda temizleme ve ayrıştırma
                    pay_payda = clean_pay_payda(row[2])
                
                    try:
                        pay, payda = pay_payda.split('/')
                        pay_num = Decimal(pay.strip())
                        payda_num = Decimal(payda.strip())
                        oran = pay_num / payda_num if payda_num != 0 else Decimal('0')
                    except (ValueError, InvalidOperation):
                        pay_num = Decimal('0')
                        payda_num = Decimal('1')
                        oran = Decimal('0')

                    # Zemin tipine göre alan değerlerini ayarla
                    if zemin_tipi in ["KatIrtifaki", "KatMulkiyeti", "KatIrtif", "KatMulki"]:
                        hisse_alan = "-"
                        toplam_alan = "-"
                    else:
                        hisse_alan = str(row[3]).strip().replace('.', ',') if row[3] else '0'
                        toplam_alan = str(row[4]).strip().replace('.', ',') if row[4] else '0'

                    # Veri satırı
                    row_data = [
                        hisse_count,
                        adi_soyadi,
                        f"'{pay_payda}",
                        int(pay_num),
                        int(payda_num),
                        f"{oran:.5f}".replace('.', ','),
                        hisse_alan,
                        toplam_alan,
                        "",
                        ""
                    ]

                    # Veri satırı formatı
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.value = value
                        cell.font = cell_font
                        cell.border = border
                        cell.alignment = alignments[chr(64 + col)]

                        # Alternatif satır renklendirmesi
                        if current_row % 2 == 0:
                            cell.fill = alt_row_fill

                        # Sayısal değerler için format
                        if col in [4, 5, 6] and isinstance(value, (int, float, Decimal)):  # Pay, Payda, Oran sütunları
                            cell.number_format = '#,##0.00'
                        elif col in [7, 8] and value != "-":  # Alan sütunları, tire olmayan değerler için
                            if isinstance(value, (int, float, Decimal)) or (isinstance(value, str) and value.replace(',', '').replace('.', '').isdigit()):
                                cell.number_format = '#,##0.00'

                    current_row += 1

                except Exception as e:
                    logging.error(f"Satır işleme hatası: {str(e)}\nVeri: {row}")
                    continue

            wb.save(file_path)
            QMessageBox.information(self, "Başarılı", "Excel dosyası oluşturuldu.")
    
        except Exception as e:
            error_msg = f"Excel dışa aktarma hatası: {str(e)}"
            logging.error(error_msg)
            QMessageBox.critical(self, "Hata", error_msg)
        finally:
            if 'conn' in locals():
                conn.close()

def main():
    app = QApplication(sys.argv)
    ex = TapuAnalyzerGUI()
    ex.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()